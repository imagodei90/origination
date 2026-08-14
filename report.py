# -*- coding: utf-8 -*-
r"""
origination — 주간 보고 생성 (HTML + 엑셀)

  python report.py [--top 20] [--band-label "자산 3천억-1조"]

HTML : 폰에서 읽는 카드형. Morning Briefing 섹션·Toolbox 탭이 같은 파일을 씀
엑셀 : 전량 랭킹 + 원시지표 + 주차 변화
"""
import os, sys, io, json, html, sqlite3, argparse
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
if (sys.stdout.encoding or "").lower().replace("-", "") != "utf8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

BASE = os.path.dirname(os.path.abspath(__file__))
DB = os.path.join(BASE, "origination.db")
OUT = os.path.join(BASE, "out")
os.makedirs(OUT, exist_ok=True)

RED = {"영업적자", "이자보상<1", "OCF음수", "자본잠식", "순차입금급증"}


def eok(v, unit="억"):
    if v is None:
        return "n.a."
    a = v / 1e8
    if abs(a) >= 10000:
        return f"{a/10000:,.1f}조"
    return f"{a:,.0f}{unit}"


def pct(v, d=1):
    return "n.a." if v is None else f"{v*100:+.{d}f}%"


def load(con, run_id=None):
    if not run_id:
        run_id = con.execute("SELECT run_id FROM tier1 ORDER BY run_id DESC LIMIT 1").fetchone()[0]
    prev = con.execute("SELECT DISTINCT run_id FROM tier1 WHERE run_id<? ORDER BY run_id DESC LIMIT 1",
                       (run_id,)).fetchone()
    prev_id = prev[0] if prev else None
    pmap = {}
    if prev_id:
        for cc, sc in con.execute("SELECT corp_code,score FROM tier1 WHERE run_id=?", (prev_id,)):
            pmap[cc] = sc
    rows = []
    q = """SELECT t.corp_code,c.corp_name,c.stock_code,t.score,t.tags,t.payload
           FROM tier1 t JOIN corp c ON c.corp_code=t.corp_code
           WHERE t.run_id=? ORDER BY t.score DESC, c.corp_name"""
    for cc, nm, sk, sc, tags, pl in con.execute(q, (run_id,)):
        p = json.loads(pl)
        rows.append({"code": cc, "name": nm, "stock": sk, "score": sc,
                     "tags": [t for t in (tags or "").split(",") if t],
                     "prev": pmap.get(cc), "new": (prev_id is not None and cc not in pmap),
                     **p})
    return run_id, prev_id, rows


def build_html(rows, run_id, prev_id, band_label, top, path):
    now = datetime.now()
    wk = now.isocalendar()
    hi = [r for r in rows if r["score"] >= 60]
    mid = [r for r in rows if 40 <= r["score"] < 60]
    newly = [r for r in rows if r["new"] and r["score"] >= 40]
    jump = [r for r in rows if r["prev"] is not None and r["score"] - r["prev"] >= 15]

    def card(r):
        badges = "".join(
            f'<span class="b {"r" if t in RED else "y"}">{html.escape(t)}</span>' for t in r["tags"])
        delta = ""
        if r["new"]:
            delta = '<span class="nw">NEW</span>'
        elif r["prev"] is not None and r["score"] != r["prev"]:
            d = r["score"] - r["prev"]
            delta = f'<span class="dl {"up" if d>0 else "dn"}">{d:+d}</span>'
        icr = "n.a." if r.get("icr") is None else f'{r["icr"]:.1f}x'
        if r.get("op_loss"):
            icr = "영업적자"
        return f"""
<div class="card">
  <div class="hd"><span class="sc s{min(r['score']//20,4)}">{r['score']}</span>
    <span class="nm">{html.escape(r['name'])}</span>
    <span class="cd">{html.escape(r['stock'] or '')}</span>{delta}</div>
  <div class="sz">자산 {eok(r.get('assets'))} · 매출 {eok(r.get('revenue'))}</div>
  <div class="gr">
    <div><i>이자보상</i><b>{icr}</b></div>
    <div><i>FCF</i><b>{eok(r.get('fcf'))}</b></div>
    <div><i>순차입금</i><b>{eok(r.get('net_debt'))}</b></div>
    <div><i>매출성장</i><b>{pct(r.get('rev_growth'))}</b></div>
    <div><i>영업마진</i><b>{pct(r.get('margin'))}</b></div>
    <div><i>OCF</i><b>{eok(r.get('ocf'))}</b></div>
  </div>
  <div class="bd">{badges}</div>
</div>"""

    def table(rs):
        tr = "".join(
            f"<tr><td>{r['score']}</td><td>{html.escape(r['name'])}</td>"
            f"<td>{html.escape(r['stock'] or '')}</td><td>{eok(r.get('assets'))}</td>"
            f"<td>{eok(r.get('revenue'))}</td><td>{pct(r.get('margin'))}</td>"
            f"<td>{html.escape(','.join(r['tags'][:3]))}</td></tr>" for r in rs)
        return (f"<table><thead><tr><th>점수</th><th>회사</th><th>코드</th><th>자산</th>"
                f"<th>매출</th><th>마진</th><th>주요신호</th></tr></thead><tbody>{tr}</tbody></table>")

    doc = f"""<!doctype html><html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Origination {wk[0]}-W{wk[1]}</title><style>
*{{box-sizing:border-box}}
body{{margin:0;padding:14px;font:15px/1.5 -apple-system,'Segoe UI',sans-serif;
background:#faf9f7;color:#1a1a1a;max-width:760px;margin:0 auto}}
h1{{font-size:19px;margin:0 0 2px}}
.sub{{color:#777;font-size:13px;margin-bottom:14px}}
.kpi{{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap}}
.kpi div{{flex:1;min-width:76px;background:#fff;border:1px solid #e6e2dc;border-radius:9px;
padding:9px;text-align:center}}
.kpi b{{display:block;font-size:20px}} .kpi i{{font-style:normal;font-size:11px;color:#888}}
h2{{font-size:14px;margin:20px 0 8px;color:#555;border-bottom:1px solid #e6e2dc;padding-bottom:5px}}
.card{{background:#fff;border:1px solid #e6e2dc;border-radius:11px;padding:12px;margin-bottom:9px}}
.hd{{display:flex;align-items:center;gap:7px;flex-wrap:wrap}}
.sc{{font-weight:700;color:#fff;border-radius:6px;padding:2px 8px;font-size:13px}}
.s0{{background:#8a9a5b}}.s1{{background:#c9a227}}.s2{{background:#d97706}}
.s3{{background:#c2410c}}.s4{{background:#991b1b}}
.nm{{font-weight:600}} .cd{{color:#999;font-size:12px}}
.nw{{background:#1d4ed8;color:#fff;font-size:10px;padding:1px 6px;border-radius:4px}}
.dl{{font-size:11px;padding:1px 6px;border-radius:4px}}
.dl.up{{background:#fee2e2;color:#991b1b}} .dl.dn{{background:#dcfce7;color:#166534}}
.sz{{color:#777;font-size:12px;margin:5px 0 8px}}
.gr{{display:grid;grid-template-columns:repeat(3,1fr);gap:7px}}
.gr div{{background:#f7f5f2;border-radius:6px;padding:6px}}
.gr i{{display:block;font-style:normal;font-size:10px;color:#888}}
.gr b{{font-size:13px}}
.bd{{margin-top:9px;display:flex;gap:5px;flex-wrap:wrap}}
.b{{font-size:10px;padding:2px 7px;border-radius:10px}}
.b.r{{background:#fee2e2;color:#991b1b}} .b.y{{background:#fef3c7;color:#92400e}}
table{{width:100%;border-collapse:collapse;font-size:12px}}
th,td{{padding:5px 6px;border-bottom:1px solid #eee;text-align:right;white-space:nowrap}}
th:nth-child(2),td:nth-child(2),th:last-child,td:last-child{{text-align:left}}
.wrap{{overflow-x:auto}}
details{{background:#fff;border:1px solid #e6e2dc;border-radius:9px;padding:10px;margin-top:8px}}
summary{{cursor:pointer;font-size:13px;font-weight:600}}
.ft{{color:#999;font-size:11px;margin-top:22px;border-top:1px solid #e6e2dc;padding-top:10px}}
@media(prefers-color-scheme:dark){{
body{{background:#16151a;color:#e8e6e3}}
.card,.kpi div,details{{background:#1f1e24;border-color:#33313a}}
.gr div{{background:#26252c}} h2{{color:#aaa;border-color:#33313a}}
th,td{{border-color:#2a2930}} .ft{{border-color:#33313a}}}}
</style></head><body>
<h1>Origination 주간 스캔</h1>
<div class="sub">{now:%Y-%m-%d} · {wk[0]}년 {wk[1]}주차 · {band_label} · {len(rows):,}사</div>
<div class="kpi">
  <div><b>{len(hi)}</b><i>고위험 60+</i></div>
  <div><b>{len(mid)}</b><i>관찰 40-59</i></div>
  <div><b>{len(newly)}</b><i>신규 진입</i></div>
  <div><b>{len(jump)}</b><i>급등 +15</i></div>
</div>
<h2>고위험 · 상위 {min(top,len(hi)) if hi else 0}사</h2>
{''.join(card(r) for r in hi[:top]) or '<div class="card">해당 없음</div>'}
<h2>관찰 대상</h2>
<div class="wrap">{table(mid[:30]) if mid else '해당 없음'}</div>
<details><summary>전체 랭킹 {len(rows):,}사 펼치기</summary>
<div class="wrap">{table(rows)}</div></details>
<div class="ft">
DART 공시 기반 자동 스캔 · 스코어는 재무 신호 조합이며 투자·자문 판단이 아님<br>
run_id {run_id}{' · 직전 ' + prev_id if prev_id else ' · 최초 실행(변화 비교 없음)'}<br>
차입금은 계정명 합산이라 일부 회사에서 미검출될 수 있음 → 태그 '차입금미확보' 확인
</div></body></html>"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(doc)
    return path


def build_xlsx(rows, run_id, prev_id, band_label, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    HF = Font(bold=True, size=9, color="FFFFFF")
    HFill = PatternFill("solid", fgColor="1F3864")
    BF = Font(size=9)
    N = '#,##0;(#,##0);"-"'
    P = '0.0%;(0.0%);"n.a."'

    def sheet(title, cols, getter):
        ws = wb.create_sheet(title)
        for i, (h, w, fmt) in enumerate(cols, 1):
            c = ws.cell(1, i, h)
            c.font, c.fill = HF, HFill
            c.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(i)].width = w
        for r_i, r in enumerate(rows, 2):
            for c_i, (h, w, fmt) in enumerate(cols, 1):
                v = getter(r, h)
                c = ws.cell(r_i, c_i, v)
                c.font = BF
                if fmt:
                    c.number_format = fmt
        ws.freeze_panes = "C2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(rows)+1}"
        return ws

    RANK = [("순위", 6, None), ("회사명", 20, None), ("종목코드", 10, None), ("스코어", 8, None),
            ("전주", 8, None), ("변화", 8, None), ("신규", 6, None), ("신호", 52, None)]
    METRIC = [("회사명", 20, None), ("종목코드", 10, None), ("스코어", 8, None),
              ("자산", 16, N), ("매출", 16, N), ("영업이익", 14, N), ("영업마진", 10, P),
              ("매출성장", 10, P), ("이자보상배율", 12, '0.00"x"'), ("순차입금", 16, N),
              ("총차입금", 16, N), ("현금성자산", 16, N), ("OCF", 16, N), ("FCF", 16, N),
              ("AR일수", 9, '0'), ("AR일수Δ", 9, '+0;-0'), ("재고일수", 9, '0'), ("재고일수Δ", 9, '+0;-0')]

    def gr(r, h):
        if h == "순위":
            return rows.index(r) + 1
        return {"회사명": r["name"], "종목코드": r["stock"], "스코어": r["score"],
                "전주": r["prev"], "변화": (r["score"] - r["prev"]) if r["prev"] is not None else None,
                "신규": "NEW" if r["new"] else "", "신호": ", ".join(r["tags"])}.get(h)

    def gm(r, h):
        m = {"회사명": r["name"], "종목코드": r["stock"], "스코어": r["score"],
             "자산": r.get("assets"), "매출": r.get("revenue"), "영업이익": r.get("ebit"),
             "영업마진": r.get("margin"), "매출성장": r.get("rev_growth"),
             "이자보상배율": r.get("icr"), "순차입금": r.get("net_debt"), "총차입금": r.get("debt"),
             "현금성자산": r.get("cash"), "OCF": r.get("ocf"), "FCF": r.get("fcf"),
             "AR일수": r.get("ar_days"), "AR일수Δ": r.get("ar_days_chg"),
             "재고일수": r.get("inv_days"), "재고일수Δ": r.get("inv_days_chg")}
        return m.get(h)

    wb.remove(wb.active)
    sheet("Ranking", RANK, gr)
    sheet("Metrics", METRIC, gm)

    ws = wb.create_sheet("Meta")
    info = [("실행 run_id", run_id), ("직전 run_id", prev_id or "최초 실행"),
            ("밴드", band_label), ("대상 회사수", len(rows)),
            ("고위험 60+", sum(1 for r in rows if r["score"] >= 60)),
            ("관찰 40-59", sum(1 for r in rows if 40 <= r["score"] < 60)),
            ("차입금 미확보", sum(1 for r in rows if not r.get("debt_found"))),
            ("생성", datetime.now().strftime("%Y-%m-%d %H:%M")),
            ("주의", "스코어는 DART 재무신호 조합. 투자·자문 판단 아님")]
    for i, (a, b) in enumerate(info, 1):
        ws.cell(i, 1, a).font = Font(bold=True, size=9)
        ws.cell(i, 2, b).font = BF
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 58
    wb.save(path)
    return path


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--top", type=int, default=20)
    ap.add_argument("--band-label", default="자산 3천억-1조")
    a = ap.parse_args()

    con = sqlite3.connect(DB)
    run_id, prev_id, rows = load(con)
    con.close()
    stamp = datetime.now().strftime("%y%m%d")
    h = build_html(rows, run_id, prev_id, a.band_label, a.top,
                   os.path.join(OUT, f"origination_{stamp}.html"))
    x = build_xlsx(rows, run_id, prev_id, a.band_label,
                   os.path.join(OUT, f"origination_{stamp}.xlsx"))
    hi = sum(1 for r in rows if r["score"] >= 60)
    mid = sum(1 for r in rows if 40 <= r["score"] < 60)
    nod = sum(1 for r in rows if not r.get("debt_found"))
    print(f"대상 {len(rows):,}사 | 고위험 {hi} | 관찰 {mid} | 차입금 미확보 {nod}")
    print(f"HTML  {h}")
    print(f"XLSX  {x}")
